from datetime import datetime
from io import BytesIO

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from models import CATEGORIES, U6_FIELD_FORMATS, Schedule


def _format_date(date_str: str) -> str:
    """Convert YYYY-MM-DD to DD/MM/YYYY, return as-is if not parseable."""
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return date_str


def _render_title_with_event(pdf, title, event_name, event_date):
    """Render page title (left, bold) with event name/date (right, italic gray) on same line."""
    y = pdf.get_y()
    x_left = pdf.l_margin
    page_w = pdf.w - pdf.l_margin - pdf.r_margin
    h = 10

    # Title on the left
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(page_w / 2, h, title)

    # Event info on the right
    if event_name or event_date:
        parts = [
            p for p in [event_name, _format_date(event_date) if event_date else ""] if p
        ]
        pdf.set_font("Helvetica", "I", 10)
        pdf.set_text_color(100, 100, 100)
        pdf.set_xy(x_left + page_w / 2, y)
        pdf.cell(page_w / 2, h, " - ".join(parts), align="R")
        pdf.set_text_color(0, 0, 0)

    pdf.set_xy(x_left, y + h)


def _render_lunch_break(pdf, total_width, sched, row_h=8):
    """Render the blue PAUSA lunch-break bar."""
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(74, 108, 247)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(
        total_width,
        row_h,
        f"PAUSA  {sched.lunch_break} min",
        border=0,
        fill=True,
        align="C",
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.set_text_color(0, 0, 0)


def _render_table_headers(pdf, col_widths, headers):
    """Render the table header row."""
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(220, 220, 220)
    col_idx = 0
    for i, h in enumerate(headers):
        if h == "Risultato (mete)":
            pdf.cell(
                col_widths[col_idx] + col_widths[col_idx + 1],
                7,
                h,
                border=1,
                fill=True,
                align="C",
            )
            col_idx += 2
        else:
            pdf.cell(col_widths[col_idx], 7, h, border=1, fill=True, align="C")
            col_idx += 1
    pdf.ln()


def _render_match_table(
    pdf, matches, sched, col_widths, headers, *, show_resting=True, show_field=True
):
    """Render the match table (header + rows + riposa + lunch break)."""
    _render_table_headers(pdf, col_widths, headers)

    row_h = 7
    total_w = sum(col_widths)
    page_bottom = pdf.h - pdf.b_margin

    resting_per_slot = sched.resting_per_slot
    match_list = list(matches)
    # Pre-compute number of matches per slot for height estimation
    slot_match_counts: dict[int, int] = {}
    for m_ in match_list:
        slot_match_counts[m_.time_slot] = slot_match_counts.get(m_.time_slot, 0) + 1

    current_slot = -1
    for idx, m in enumerate(match_list):
        is_first_in_slot = idx == 0 or match_list[idx - 1].time_slot != m.time_slot

        if is_first_in_slot:
            # Render riposa for the previous slot
            if current_slot >= 0 and show_resting:
                resting = resting_per_slot.get(current_slot, [])
                if resting:
                    pdf.set_font("Helvetica", "I", 7)
                    pdf.set_fill_color(248, 248, 248)
                    pdf.cell(
                        sum(col_widths),
                        4,
                        f"Riposa: {', '.join(resting)}",
                        border=1,
                        fill=True,
                        new_x="LMARGIN",
                        new_y="NEXT",
                    )

            # Lunch break
            if sched.morning_slots > 0 and m.time_slot == sched.morning_slots:
                if pdf.get_y() + 10 > page_bottom:
                    pdf.add_page()
                    _render_table_headers(pdf, col_widths, headers)
                _render_lunch_break(pdf, sum(col_widths), sched)

            # Check if the entire new slot fits on the current page
            slot_count = slot_match_counts[m.time_slot]
            has_resting = bool(resting_per_slot.get(m.time_slot))
            slot_height = slot_count * row_h + (
                4 if show_resting and has_resting else 0
            )
            if pdf.get_y() + slot_height > page_bottom:
                pdf.add_page()
                _render_table_headers(pdf, col_widths, headers)
                is_first_in_slot = True

            current_slot = m.time_slot

        is_last_in_slot = (
            idx == len(match_list) - 1 or match_list[idx + 1].time_slot != m.time_slot
        )

        y_top = pdf.get_y()
        x_left = pdf.l_margin

        # Draw cells without borders
        pdf.set_font("Helvetica", "", 8)
        ci = 0
        pdf.cell(col_widths[ci], row_h, m.start_time, border=0, align="C")
        ci += 1
        if show_field:
            pdf.cell(
                col_widths[ci], row_h, f"Campo {m.field_number}", border=0, align="C"
            )
            ci += 1
        pdf.cell(col_widths[ci], row_h, m.team1.upper(), border=0)
        ci += 1
        pdf.cell(col_widths[ci], row_h, m.team2.upper(), border=0)
        ci += 1
        pdf.cell(col_widths[ci], row_h, "", border=0, align="C")
        ci += 1
        pdf.cell(col_widths[ci], row_h, "", border=0, align="C")
        ci += 1
        if not sched.no_referee:
            pdf.cell(col_widths[ci], row_h, m.referee.upper(), border=0, align="C")
        pdf.ln()

        y_bottom = y_top + row_h

        # Horizontal lines: black at slot boundaries, light gray inside
        pdf.set_draw_color(0, 0, 0)
        if is_first_in_slot:
            pdf.line(x_left, y_top, x_left + total_w, y_top)
        else:
            pdf.set_draw_color(200, 200, 200)
            pdf.line(x_left, y_top, x_left + total_w, y_top)
            pdf.set_draw_color(0, 0, 0)
        if is_last_in_slot:
            pdf.line(x_left, y_bottom, x_left + total_w, y_bottom)
        else:
            pdf.set_draw_color(200, 200, 200)
            pdf.line(x_left, y_bottom, x_left + total_w, y_bottom)
            pdf.set_draw_color(0, 0, 0)

        # Vertical lines: black for outer edges, light gray for internal
        pdf.line(x_left, y_top, x_left, y_bottom)
        pdf.set_draw_color(200, 200, 200)
        x = x_left
        for cw in col_widths[:-1]:
            x += cw
            pdf.line(x, y_top, x, y_bottom)
        pdf.set_draw_color(0, 0, 0)
        pdf.line(x_left + total_w, y_top, x_left + total_w, y_bottom)

    # Riposa row for the last slot
    if show_resting and current_slot >= 0:
        resting = resting_per_slot.get(current_slot, [])
        if resting:
            pdf.set_font("Helvetica", "I", 7)
            pdf.set_fill_color(248, 248, 248)
            pdf.cell(
                sum(col_widths),
                4,
                f"Riposa: {', '.join(resting)}",
                border=1,
                fill=True,
                new_x="LMARGIN",
                new_y="NEXT",
            )


def _render_team_page(pdf, team_name, sched, event_name, event_date):
    """Render a per-team page showing their chronological activity."""
    pdf.add_page()
    _render_title_with_event(
        pdf, f"{sched.category} - {team_name.upper()}", event_name, event_date
    )
    pdf.ln(2)

    # Table
    col_widths = [30, 25, 30, 105]
    headers = ["Orario", "Campo", "Attivita", "Dettagli"]

    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(220, 220, 220)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, h, border=1, fill=True, align="C")
    pdf.ln()

    slots = sorted({m.time_slot for m in sched.matches})
    resting_per_slot = sched.resting_per_slot

    for slot in slots:
        # Lunch break
        if sched.morning_slots > 0 and slot == sched.morning_slots:
            _render_lunch_break(pdf, sum(col_widths), sched)

        slot_matches = [m for m in sched.matches if m.time_slot == slot]
        start_time = slot_matches[0].start_time

        activity = ""
        campo = ""
        details = ""

        for m in slot_matches:
            if team_name in (m.team1, m.team2):
                activity = "GIOCA"
                campo = str(m.field_number)
                opponent = m.team2 if team_name == m.team1 else m.team1
                details = f"vs {opponent.upper()}"
                break
            if not sched.no_referee and m.referee == team_name:
                activity = "ARBITRA"
                campo = str(m.field_number)
                details = f"{m.team1.upper()} vs {m.team2.upper()}"
                break

        if not activity and team_name in resting_per_slot.get(slot, []):
            activity = "RIPOSA"

        if not activity:
            continue

        pdf.set_font("Helvetica", "", 9)
        pdf.cell(col_widths[0], 8, start_time, border=1, align="C")
        pdf.cell(col_widths[1], 8, campo, border=1, align="C")
        pdf.cell(col_widths[2], 8, activity, border=1, align="C")
        pdf.cell(col_widths[3], 8, details, border=1)
        pdf.ln()


def schedule_to_pdf(
    schedules: list[Schedule],
    event_name: str = "",
    event_date: str = "",
    *,
    include_main: bool = True,
    include_field: bool = True,
    include_team: bool = True,
) -> bytes:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=10)

    for sched in schedules:
        config = CATEGORIES[sched.category]

        # Column widths needed by both main and field pages
        if sched.no_referee:
            col_widths = [15, 20, 60, 60, 18, 18]
            headers = ["Orario", "Campo", "Squadra A", "Squadra B", "Risultato (mete)"]
        else:
            col_widths = [15, 20, 40, 40, 18, 18, 40]
            headers = [
                "Orario",
                "Campo",
                "Squadra A",
                "Squadra B",
                "Risultato (mete)",
                "Arbitro",
            ]

        if include_main:
            pdf.add_page()
            _render_title_with_event(
                pdf, f"Calendario {sched.category}", event_name, event_date
            )
            pdf.set_font("Helvetica", "", 10)
            if sched.half_time_interval > 0:
                half = sched.match_duration // 2
                match_desc = f"{half} min + {half} min ({sched.half_time_interval} min intervallo)"
            else:
                match_desc = f"{sched.match_duration} min"
            pdf.cell(
                0,
                6,
                f"{len(sched.stats)} squadre | {len(sched.matches)} partite | "
                f"{match_desc} + {sched.break_duration} min pausa",
                new_x="LMARGIN",
                new_y="NEXT",
            )
            pdf.ln(2)

            # Warnings
            for w in sched.warnings:
                pdf.set_font("Helvetica", "I", 9)
                pdf.cell(0, 5, f"Nota: {w}", new_x="LMARGIN", new_y="NEXT")
            if sched.warnings:
                pdf.ln(1)

            _render_match_table(pdf, sched.matches, sched, col_widths, headers)

            pdf.ln(5)

            # --- Stats table (left) + Field diagram (right) side by side ---
            stats_table_w = 95
            gap = 10
            draw_w = 80
            draw_h = draw_w * config.field_width_max / config.field_length_max
            meta_mm = draw_w * config.meta / config.field_length_max

            num_teams = len(sched.stats)
            stats_h = 8 + 7 + num_teams * 6
            section_h = max(stats_h, draw_h + 18) + 5
            if pdf.get_y() + section_h > pdf.h - pdf.b_margin:
                pdf.add_page()

            section_top = pdf.get_y()

            # -- Stats table (left column) --
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(stats_table_w, 8, "Statistiche", new_x="LMARGIN", new_y="NEXT")

            if sched.no_referee:
                stat_widths = [55, 18, 22]
                stat_headers = ["Squadra", "Partite", "Max Attesa"]
            else:
                stat_widths = [35, 18, 20, 22]
                stat_headers = ["Squadra", "Partite", "Arbitraggi", "Max Attesa"]

            pdf.set_font("Helvetica", "B", 9)
            pdf.set_fill_color(220, 220, 220)
            for i, h in enumerate(stat_headers):
                pdf.cell(stat_widths[i], 7, h, border=1, fill=True)
            pdf.ln()

            pdf.set_font("Helvetica", "", 8)
            for team, stat in sched.stats.items():
                pdf.cell(stat_widths[0], 6, team, border=1)
                pdf.cell(stat_widths[1], 6, str(stat["played"]), border=1)
                if not sched.no_referee:
                    pdf.cell(stat_widths[2], 6, str(stat["refereed"]), border=1)
                    pdf.cell(stat_widths[3], 6, f"{stat['max_wait']} min", border=1)
                else:
                    pdf.cell(stat_widths[2], 6, f"{stat['max_wait']} min", border=1)
                pdf.ln()

            stats_bottom = pdf.get_y()

            # -- Field diagram (right column) --
            field_x = pdf.l_margin + stats_table_w + gap
            field_y = section_top

            pdf.set_font("Helvetica", "B", 11)
            pdf.set_xy(field_x, field_y)
            pdf.cell(draw_w, 8, "Dimensioni Campo")
            field_y += 8

            pdf.set_fill_color(144, 190, 109)
            pdf.set_draw_color(50, 100, 50)
            pdf.set_line_width(0.5)
            pdf.rect(field_x, field_y, draw_w, draw_h, style="DF")

            pdf.set_draw_color(255, 255, 255)
            pdf.set_line_width(0.3)
            pdf.dashed_line(
                field_x + meta_mm,
                field_y,
                field_x + meta_mm,
                field_y + draw_h,
                dash_length=2,
                space_length=1.5,
            )
            pdf.dashed_line(
                field_x + draw_w - meta_mm,
                field_y,
                field_x + draw_w - meta_mm,
                field_y + draw_h,
                dash_length=2,
                space_length=1.5,
            )

            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 7)
            pdf.set_xy(field_x + 1, field_y + draw_h / 2 - 2)
            pdf.cell(meta_mm - 2, 4, f"{config.meta}m", align="C")
            pdf.set_xy(field_x + draw_w - meta_mm + 1, field_y + draw_h / 2 - 2)
            pdf.cell(meta_mm - 2, 4, f"{config.meta}m", align="C")

            pdf.set_text_color(0, 0, 0)
            pdf.set_draw_color(0, 0, 0)

            pdf.set_font("Helvetica", "", 8)
            pdf.set_xy(field_x, field_y + draw_h + 1)
            pdf.cell(draw_w, 4, config.field_length, align="C")
            pdf.set_xy(field_x + draw_w + 2, field_y + draw_h / 2 - 2)
            pdf.cell(20, 4, config.field_width)

            # For U6, list all 3 field formats below the diagram
            if sched.category == "U6":
                pdf.set_font("Helvetica", "B", 8)
                pdf.set_xy(field_x, field_y + draw_h + 6)
                pdf.cell(
                    draw_w,
                    5,
                    "Dimensioni variabili in base al numero di giocatori:",
                    new_x="LMARGIN",
                    new_y="NEXT",
                )
                pdf.set_font("Helvetica", "", 8)
                for players, fmt in U6_FIELD_FORMATS.items():
                    pdf.set_x(field_x)
                    pdf.cell(
                        draw_w,
                        4,
                        f"{players} vs {players}:  {fmt['field_width']} \u00d7 {fmt['field_length']}",
                        new_x="LMARGIN",
                        new_y="NEXT",
                    )

            pdf.set_y(max(stats_bottom, field_y + draw_h + 8) + 4)

        # --- Per-field pages ---
        if include_field:
            if sched.no_referee:
                field_col_widths = [15, 65, 65, 18, 18]
                field_headers = ["Orario", "Squadra A", "Squadra B", "Risultato (mete)"]
            else:
                field_col_widths = [15, 45, 45, 18, 18, 50]
                field_headers = [
                    "Orario",
                    "Squadra A",
                    "Squadra B",
                    "Risultato (mete)",
                    "Arbitro",
                ]

            field_numbers = sorted({m.field_number for m in sched.matches})
            for fn in field_numbers:
                pdf.add_page()
                _render_title_with_event(
                    pdf, f"{sched.category} - Campo {fn}", event_name, event_date
                )
                pdf.ln(2)

                field_matches = [m for m in sched.matches if m.field_number == fn]
                _render_match_table(
                    pdf,
                    field_matches,
                    sched,
                    field_col_widths,
                    field_headers,
                    show_resting=False,
                    show_field=False,
                )

        # --- Per-team pages ---
        if include_team:
            for team_name in sched.stats:
                _render_team_page(pdf, team_name, sched, event_name, event_date)

    return pdf.output()


def schedule_to_excel(
    schedules: list[Schedule], event_name: str = "", event_date: str = ""
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
    )

    for sched in schedules:
        ws = wb.create_sheet(title=sched.category)

        # Event header
        if event_name or event_date:
            header_parts = [
                p
                for p in [event_name, _format_date(event_date) if event_date else ""]
                if p
            ]
            ws.append([" - ".join(header_parts)])
            ws.cell(ws.max_row, 1).font = Font(bold=True, size=13)
            ws.append([])

        # Schedule table
        if sched.no_referee:
            headers = ["#", "Orario", "Campo", "Squadra 1", "Squadra 2"]
        else:
            headers = ["#", "Orario", "Campo", "Squadra 1", "Squadra 2", "Arbitro"]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill

        resting_per_slot_xl = sched.resting_per_slot
        current_slot_xl = -1
        num_cols = len(headers)
        lunch_fill = PatternFill(
            start_color="4A6CF7", end_color="4A6CF7", fill_type="solid"
        )
        for m in sched.matches:
            if m.time_slot != current_slot_xl:
                if current_slot_xl >= 0:
                    resting = resting_per_slot_xl.get(current_slot_xl, [])
                    if resting:
                        row = [f"Riposa: {', '.join(resting)}"] + [""] * (num_cols - 1)
                        ws.append(row)
                        ws.cell(ws.max_row, 1).font = Font(italic=True, color="888888")
                # Lunch break row
                if sched.morning_slots > 0 and m.time_slot == sched.morning_slots:
                    row = [f"PAUSA  {sched.lunch_break} min"] + [""] * (num_cols - 1)
                    ws.append(row)
                    for cell in ws[ws.max_row]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = lunch_fill
                current_slot_xl = m.time_slot
            row = [
                m.match_number,
                m.start_time,
                f"Campo {m.field_number}",
                m.team1,
                m.team2,
            ]
            if not sched.no_referee:
                row.append(m.referee)
            ws.append(row)

        # Riposa for the last slot
        if current_slot_xl >= 0:
            resting = resting_per_slot_xl.get(current_slot_xl, [])
            if resting:
                row = [f"Riposa: {', '.join(resting)}"] + [""] * (num_cols - 1)
                ws.append(row)
                ws.cell(ws.max_row, 1).font = Font(italic=True, color="888888")

        # Blank row then stats
        ws.append([])
        stats_start = ws.max_row + 1
        if sched.no_referee:
            ws.append(["Squadra", "Partite Giocate", "Max Attesa (min)"])
        else:
            ws.append(["Squadra", "Partite Giocate", "Arbitraggi", "Max Attesa (min)"])
        for cell in ws[stats_start]:
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill

        for team, stat in sched.stats.items():
            if sched.no_referee:
                ws.append([team, stat["played"], stat["max_wait"]])
            else:
                ws.append([team, stat["played"], stat["refereed"], stat["max_wait"]])

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
